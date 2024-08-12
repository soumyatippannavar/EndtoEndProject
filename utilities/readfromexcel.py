from openpyxl import Workbook, load_workbook


def getRow(file,sheetname):
    wb = load_workbook(file)
    sh = wb[sheetname]
    return (sh.max_row)


def getCol(file, sheetname):
    wb = load_workbook(file)
    sh = wb[sheetname]
    return (sh.max_column)


def readdata(file,sheetname,rownum,colnum):
    wb = load_workbook(file)
    sh = wb[sheetname]
    return (sh.cell(row = rownum, column = colnum).value)