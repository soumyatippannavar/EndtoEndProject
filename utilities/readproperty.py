from ddt import ddt , data, unpack
from openpyxl.reader.excel import load_workbook


#class readproperty:
def read_data(filename,sheet):
    data_list=[]
    wr = load_workbook(filename=filename)  # reading data from excel
    sh = wr[sheet]
    #print(sh['A1'].value)
    row_cont = sh.max_row
    col_cont = sh.max_column
    #print(row_cont)
    #print(col_cont)

    for i in range(2,row_cont):
        row=[]
        for j in range(1,col_cont):
            row.append(sh.cell(row=i,column=j).value)
            data_list.append(row)
    return data_list





